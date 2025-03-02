import jieba
from openpyxl import load_workbook, Workbook


# 定义一个函数来读取Excel文件中的标题和内容
def read_titles_and_contents(excel_path):
    wb = load_workbook(excel_path)
    sheet = wb.active
    titles_and_contents = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        title, url, content = row[1], row[2], row[3]
        if title and content:  # 确保标题和内容不为空
            titles_and_contents.append((title, content))
    return titles_and_contents


# 定义一个函数将分词结果写入Excel的指定工作表
def write_to_excel(wb, title, segmented_text, stopwords):
    # 去除停用词和单字词
    filtered_text = ' '.join([word for word in segmented_text if word not in stopwords and len(word) > 1])

    # 写入表头如果还没有写入
    if '内容' not in wb.sheetnames:
        ws = wb.create_sheet('内容')
        ws.append(["标题", "分词后内容"])
    else:
        ws = wb['内容']

    # 将分词结果写入工作表
    ws.append([title, filtered_text])


# 主函数
if __name__ == "__main__":
    excel_path = '史记章节.xlsx'  # 这里填写你的Excel文件路径
    titles_and_contents = read_titles_and_contents(excel_path)

    # 读入停用词档，并去除停用词前后空白、换行符
    with open('古文停用词表.txt', "r", encoding='UTF-8') as file:
        stopwords = file.readlines()
    stopwords = [s.strip() for s in stopwords]

    # 加载现有的Excel工作簿
    wb = load_workbook(excel_path)

    # 遍历标题和内容，分词并写入Excel工作表
    for title, content in titles_and_contents:
        segmented_text = jieba.cut(content)  # 对内容进行分词
        write_to_excel(wb, title, segmented_text, stopwords)  # 写入Excel工作表

    # 保存新的Excel文件
    new_excel_path = '史记分词后.xlsx'
    wb.save(new_excel_path)
    print(f"分词结果已成功写入文件：{new_excel_path}")